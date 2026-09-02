"""
Mestre do Pano — sync_stock.py (Fase 6.5)

Lê o Stock.xlsx (folha "Rascunhos") e gera data/products.json + copia as
fotografias associadas a cada produto para images/produtos/<id>/.

O Stock.xlsx NUNCA é alterado por este script — é apenas lido.

Novidade da Fase 6.5 — AGRUPAMENTO DE VARIAÇÕES:
    O Stock.xlsx não tem um sistema de variantes nativo — cada cor/pack é
    uma linha própria. As colunas "agrupável" (sim/não), "tipo" (cor |
    quantidade | -) e "variação" (ex.: "Verde", 6, 18) permitem agrupar
    linhas que são, na prática, a mesma peça em versões diferentes, num
    único "produto pai" com uma lista `variations: []`.

    Regras:
      - agrupável == "não" (ou em falta) → produto simples, sem `variations`
        (mantém a forma da Fase 5: campos no topo do objeto).
      - agrupável == "sim" → a linha entra num grupo com outras linhas do
        mesmo `tipo` cujo título, depois de remover as palavras da própria
        variação/cor, seja idêntico. O grupo produz UM objeto no JSON com
        `variations: [...]` — nunca dois produtos "pai" para a mesma peça.
      - Nunca inventamos uma variação: se `variacao` estiver vazia/`-`
        numa linha marcada `agrupável == sim`, essa linha é tratada como
        produto simples e um aviso é emitido (dados incompletos no Excel,
        não avariados pelo script).

A coluna "estado" (E) é lida do Excel apenas para detetar a sua presença
sem quebrar o mapeamento de colunas, mas é EXCLUÍDA deliberadamente do
products.json — não faz parte do schema desta fase.

Uso:
    python sync_stock.py

Configuração: ver config.json (copiar de config.example.json e editar).
"""

import io
import json
import os
import re
import shutil
import sys
import unicodedata
from pathlib import Path

# Windows (cp1252) nao consegue imprimir emojis (ex.: nos titulos de
# produtos copiados do Excel) - forcar stdout/stderr para UTF-8 evita o
# UnicodeEncodeError ao imprimir avisos/nomes de produtos com emoji.
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
if sys.stderr.encoding and sys.stderr.encoding.lower() != "utf-8":
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent
CONFIG_PATH = BASE_DIR / "config.json"

# Colunas obrigatórias para um produto ser considerado válido.
CAMPOS_OBRIGATORIOS = ["titulo", "preco", "stock"]

# Colunas do Excel que são internas e NUNCA devem ir para o products.json.
# "estado" foi movido para aqui na Fase 6.5 — deixa de ser exposto no JSON
# (era usado para mostrar "Estado: Novo com etiquetas" na página de
# produto, UI removida nesta fase).
CAMPOS_INTERNOS = {"custo", "estado"}

EXTENSOES_IMAGEM = {".jpg", ".jpeg", ".png", ".webp"}

# --- Peso (coluna real do Stock.xlsx: "peso_kg", em QUILOGRAMAS) --------
COLUNA_PESO = "peso_kg"

# --- Unidades (coluna real do Stock.xlsx: "unidades") ---------------------
COLUNA_UNIDADES = "unidades"

# --- Agrupamento de variações (Fase 6.5): colunas O, P, Q do Stock.xlsx --
COLUNA_AGRUPAVEL = "agrupável"
COLUNA_TIPO = "tipo"
COLUNA_VARIACAO = "variação"

VALORES_AGRUPAVEL_SIM = {"sim", "yes", "true", "1"}
TIPOS_VARIACAO_VALIDOS = {"cor", "quantidade"}

# Palavras de cor conhecidas (nomes + modificadores comuns em PT) usadas
# para (a) tornar o agrupamento robusto a pequenas diferenças entre o
# título e o texto exato da coluna "variação"/"cor" (ex.: título diz
# "Azul Escuro", a coluna variação diz "Azul marinho" — ambos são "cor"
# e devem cair no mesmo grupo do mesmo produto-base) e (b) limpar o nome
# de exibição do grupo. Lista fixa e documentada — não é deteção mágica,
# é só uma normalização de texto.
PALAVRAS_COR_NORMALIZACAO = {
    "verde", "lilas", "cinzento", "castanho", "azul", "amarelo", "rosa",
    "vermelho", "bordeaux", "vinho", "turquesa", "multi", "branco", "preto",
    "cinza", "laranja", "roxo", "marinho", "escuro", "escura", "claro",
    "clara", "agua", "forte",
}


def resolver_unit_count(dados: dict):
    """
    Lê exclusivamente a coluna 'Unidades' do Excel.

    Devolve (unit_count: int|None, motivo_invalido: str|None).
    unit_count só é devolvido quando o valor é um número inteiro > 0.
    Nunca assume 1 por omissão — um produto sem 'Unidades' válido fica
    identificado num aviso e o campo fica a null no products.json.
    """
    valor = dados.get(COLUNA_UNIDADES)

    if valor is None or str(valor).strip() == "":
        return None, "sem valor"

    texto = str(valor).strip()

    try:
        numero = float(texto.replace(",", "."))
    except (TypeError, ValueError):
        return None, f"valor não numérico ({texto!r})"

    if numero != int(numero):
        return None, f"valor não é um número inteiro ({texto!r})"

    unit_count = int(numero)

    if unit_count <= 0:
        return None, f"valor inválido (<= 0): {texto!r}"

    return unit_count, None


def resolver_peso_g(dados: dict):
    """
    Lê exclusivamente a coluna 'peso_kg' do Excel (valor em QUILOGRAMAS,
    ex.: 0,3 = 300 g) e converte para GRAMAS (weight_g), a unidade usada
    internamente em todo o site.

    Devolve (weight_g: int|None, motivo_invalido: str|None).
    weight_g só é devolvido quando o valor é numérico e > 0.
    """
    valor = dados.get(COLUNA_PESO)

    if valor is None or str(valor).strip() == "":
        return None, "sem valor"

    texto = str(valor).strip()

    try:
        peso_kg = float(texto.replace(",", "."))
    except (TypeError, ValueError):
        return None, f"valor não numérico ({texto!r})"

    peso_g = round(peso_kg * 1000)

    if peso_g <= 0:
        return None, f"valor inválido (<= 0): {texto!r}"

    return peso_g, None


def carregar_config():
    if not CONFIG_PATH.exists():
        print("ERRO: config.json não encontrado.")
        print(f"Copia 'config.example.json' para 'config.json' em {BASE_DIR} e edita os caminhos.")
        sys.exit(1)
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def verificar_ficheiro_disponivel_localmente(caminho: Path):
    """Deteta o caso do OneDrive Files On-Demand (ficheiro só na cloud, não local)."""
    if not caminho.exists():
        print("ERRO:")
        print(f"Stock.xlsx não foi encontrado em: {caminho}")
        print("Confirma o caminho em config.json.")
        sys.exit(1)
    try:
        tamanho = caminho.stat().st_size
        with open(caminho, "rb") as f:
            inicio = f.read(4)
        if tamanho == 0 or not inicio:
            raise OSError("ficheiro vazio ou inacessível")
    except OSError:
        print("ERRO:")
        print("Stock.xlsx não está disponível localmente.")
        print("Abra/sincronize o ficheiro através do OneDrive e tente novamente.")
        sys.exit(1)


def slugify(texto: str) -> str:
    """Gera um id simples e estável a partir do título (sem SKU no Excel)."""
    texto = unicodedata.normalize("NFKD", texto)
    texto = texto.encode("ascii", "ignore").decode("ascii")  # remove acentos e emojis
    texto = texto.lower()
    texto = re.sub(r"[^a-z0-9]+", "-", texto)
    return texto.strip("-")


def normalizar_pasta_fotos(pasta_relativa: str) -> str:
    """Converte separadores Windows ('\\') para barras Web ('/') — nunca
    grava um caminho com '\\' no products.json, mesmo que o Excel os use."""
    return pasta_relativa.replace("\\", "/").strip()


def _normalizar_texto_para_chave(texto: str) -> str:
    """Normalização ASCII/minúsculas usada só para calcular chaves de
    agrupamento — nunca é usada para texto mostrado ao utilizador."""
    texto = unicodedata.normalize("NFKD", str(texto))
    texto = texto.encode("ascii", "ignore").decode("ascii")
    texto = texto.lower()
    texto = re.sub(r"[^a-z0-9 ]", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def calcular_chave_grupo(titulo: str, cor: str, tipo: str, variacao) -> str:
    """
    Deriva uma chave de agrupamento a partir do título, removendo os
    tokens da variação (e, quando tipo=='cor', também os nomes de cor
    conhecidos) — para que "Pareô de Praia Verde" e "Pareô de Praia
    Lilás" caiam na mesma chave, mas "Conjunto 6 Panos de Favo Vermelho"
    e "Conjunto 6 Panos de Favo Azul" NÃO caiam (a cor faz parte da
    identidade do produto quando o tipo de variação é 'quantidade').
    """
    tokens = _normalizar_texto_para_chave(titulo).split(" ")

    tokens_variacao = set(_normalizar_texto_para_chave(variacao).split(" ")) if variacao else set()
    tokens = [t for t in tokens if t not in tokens_variacao]

    if tipo == "cor":
        tokens = [t for t in tokens if t not in PALAVRAS_COR_NORMALIZACAO]
        base = " ".join(t for t in tokens if t)
        return f"cor::{base}"

    if tipo == "quantidade":
        # a cor faz parte da identidade do grupo (Favo Vermelho != Favo Azul)
        tokens_cor = _normalizar_texto_para_chave(cor).split(" ") if cor else []
        tokens = [t for t in tokens if not t.isdigit()]
        base = " ".join(t for t in tokens if t)
        cor_chave = "_".join(sorted(set(tokens_cor)))
        return f"quantidade::{base}::{cor_chave}"

    # nunca deveria chegar aqui (chamado só quando tipo é válido)
    return f"outro::{_normalizar_texto_para_chave(titulo)}"


def calcular_nome_grupo(titulo_original: str, cor: str, tipo: str, variacao) -> str:
    """
    Remove do título ORIGINAL (com acentos/maiúsculas preservados) os
    tokens da variação/cor, para obter um nome de exibição limpo e comum
    a todas as linhas do grupo (ex.: 'Pareô de Praia - Toalha Versátil
    para Viagem e Praia', sem a cor).
    """
    texto = titulo_original
    tokens_a_remover = set()

    if variacao not in (None, "", "-"):
        tokens_a_remover.update(str(variacao).split())

    if tipo == "cor":
        if cor:
            tokens_a_remover.update(str(cor).replace("-", " ").split())
        tokens_a_remover.update(PALAVRAS_COR_NORMALIZACAO)
        tokens_a_remover.update(w.capitalize() for w in PALAVRAS_COR_NORMALIZACAO)

    for token in tokens_a_remover:
        token = token.strip()
        if not token:
            continue
        texto = re.sub(rf"\b{re.escape(token)}\b", "", texto, flags=re.IGNORECASE)

    texto = re.sub(r"\s{2,}", " ", texto)
    texto = re.sub(r"\s*-\s*-\s*", " - ", texto)
    texto = texto.strip(" -").strip()
    return texto or titulo_original


def resolver_agrupamento(dados: dict):
    """
    Lê as colunas 'agrupável', 'tipo' e 'variação' de uma linha.

    Devolve (agrupavel: bool, tipo: str|None, variacao_valor|None, aviso: str|None).
    Nunca assume agrupável=True por omissão — só quando a coluna diz
    explicitamente "sim" E existe um tipo/variação válidos.
    """
    valor_agrupavel = str(dados.get(COLUNA_AGRUPAVEL) or "").strip().lower()
    agrupavel = valor_agrupavel in VALORES_AGRUPAVEL_SIM

    if not agrupavel:
        return False, None, None, None

    tipo = str(dados.get(COLUNA_TIPO) or "").strip().lower()
    variacao_bruta = dados.get(COLUNA_VARIACAO)
    variacao_texto = str(variacao_bruta).strip() if variacao_bruta is not None else ""

    if tipo not in TIPOS_VARIACAO_VALIDOS:
        return False, None, None, (
            f"marcado como agrupável='sim' mas 'tipo' ({tipo!r}) não é "
            f"'cor' nem 'quantidade' — tratado como produto simples."
        )

    if not variacao_texto or variacao_texto == "-":
        return False, None, None, (
            f"marcado como agrupável='sim' e tipo={tipo!r} mas 'variação' "
            f"está vazia/'-' — tratado como produto simples (corrigir no Excel)."
        )

    # normaliza variação de quantidade para número quando possível (ex.: 6, 18)
    if tipo == "quantidade":
        try:
            variacao_valor = int(float(str(variacao_bruta).replace(",", ".")))
        except (TypeError, ValueError):
            variacao_valor = variacao_texto
    else:
        variacao_valor = variacao_texto

    return True, tipo, variacao_valor, None


def ler_produtos(caminho_excel: Path, nome_folha: str):
    wb = openpyxl.load_workbook(caminho_excel, data_only=True)
    if nome_folha not in wb.sheetnames:
        print(f"ERRO: a folha '{nome_folha}' não existe no Excel. Folhas encontradas: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[nome_folha]

    linhas = list(ws.iter_rows(values_only=True))
    cabecalho = [str(c).strip() if c else "" for c in linhas[0]]

    linhas_lidas = []
    avisos = []
    erros = []

    for num_linha, linha in enumerate(linhas[1:], start=2):
        dados = dict(zip(cabecalho, linha))
        if all(v is None or str(v).strip() == "" for v in dados.values()):
            continue  # linha vazia, ignorar

        campos_em_falta = [
            campo for campo in CAMPOS_OBRIGATORIOS
            if dados.get(campo) is None or str(dados.get(campo)).strip() == ""
        ]
        if campos_em_falta:
            erros.append(f"Produto na linha {num_linha} não possui: {', '.join(campos_em_falta)}.")
            continue

        titulo = str(dados["titulo"]).strip()
        titulo = re.sub(r"\s+", " ", titulo)

        categoria = (str(dados.get("categoria")) or "").strip() or None
        cor = (str(dados.get("cor")) or "").strip() or None

        weight_g, motivo_invalido = resolver_peso_g(dados)
        if weight_g is None:
            avisos.append(
                f"WARNING: Produto '{titulo}' (linha {num_linha}) não possui "
                f"'{COLUNA_PESO}' válido no Stock.xlsx ({motivo_invalido}). "
                f"weight_g ficará a null."
            )

        unit_count, motivo_unidades_invalido = resolver_unit_count(dados)
        if unit_count is None:
            avisos.append(
                f"WARNING: Produto '{titulo}' (linha {num_linha}) não possui "
                f"'{COLUNA_UNIDADES}' válido no Stock.xlsx ({motivo_unidades_invalido}). "
                f"unit_count ficará a null."
            )

        agrupavel, tipo_variacao, variacao_valor, aviso_agrupamento = resolver_agrupamento(dados)
        if aviso_agrupamento:
            avisos.append(f"WARNING: Produto '{titulo}' (linha {num_linha}) — {aviso_agrupamento}")

        pasta_fotos = normalizar_pasta_fotos(str(dados.get("pasta_fotos") or "").strip())

        linha_produto = {
            "titulo": titulo,
            "description": (str(dados.get("descricao")) or "").strip(),
            "price": round(float(dados["preco"]), 2),
            "brand": (str(dados.get("marca")) or "").strip() or None,
            "category": categoria,
            "color": cor,
            "material": (str(dados.get("material")) or "").strip() or None,
            "additional_info": (str(dados.get("mensagem_personalizada")) or "").strip() or None,
            "stock": int(dados["stock"]),
            "weight_g": weight_g,
            "unit_count": unit_count,
            "_pasta_fotos": pasta_fotos,
            "_agrupavel": agrupavel,
            "_tipo_variacao": tipo_variacao,
            "_variacao_valor": variacao_valor,
            "_num_linha": num_linha,
        }

        # nunca deixar passar campos internos, mesmo que o nome mude no futuro
        for campo_interno in CAMPOS_INTERNOS:
            linha_produto.pop(campo_interno, None)

        linhas_lidas.append(linha_produto)

    return linhas_lidas, avisos, erros


def agrupar_produtos(linhas_lidas, avisos):
    """
    Agrupa linhas marcadas como agrupável=True (mesmo 'tipo' + chave de
    título comum) em produtos-pai com `variations: []`; linhas não
    agrupáveis tornam-se produtos simples (schema igual à Fase 5).

    Devolve a lista final de produtos (cada um já com 'id' definitivo).
    """
    grupos = {}  # chave -> lista de linhas (na ordem em que aparecem no Excel)
    ordem_grupos = []
    produtos_simples = []

    for linha in linhas_lidas:
        if linha["_agrupavel"]:
            chave = calcular_chave_grupo(
                linha["titulo"], linha["color"], linha["_tipo_variacao"], linha["_variacao_valor"]
            )
            if chave not in grupos:
                grupos[chave] = []
                ordem_grupos.append(chave)
            grupos[chave].append(linha)
        else:
            produtos_simples.append(linha)

    produtos_finais = []
    ids_usados = {}

    def gerar_id(base_texto):
        base_id = slugify(base_texto) or "produto"
        contador = ids_usados.get(base_id, 0)
        id_final = base_id if not contador else f"{base_id}-{contador + 1}"
        ids_usados[base_id] = contador + 1
        return id_final

    # produtos simples primeiro, para preservar a mesma ordem/ids da Fase 5
    # quando não há agrupamento nenhum.
    for linha in linhas_lidas:
        pass  # ordem final é reconstituída abaixo, respeitando a ordem do Excel

    # Reconstituir a ordem original do Excel: percorremos linhas_lidas e,
    # para cada grupo, só emitimos o produto-pai na primeira ocorrência.
    grupos_emitidos = set()
    for linha in linhas_lidas:
        if not linha["_agrupavel"]:
            id_final = gerar_id(linha["titulo"])
            produto = {
                "id": id_final,
                "name": linha["titulo"],
                "description": linha["description"],
                "price": linha["price"],
                "brand": linha["brand"],
                "category": linha["category"],
                "color": linha["color"],
                "material": linha["material"],
                "additional_info": linha["additional_info"],
                "stock": linha["stock"],
                "weight_g": linha["weight_g"],
                "unit_count": linha["unit_count"],
                "images": [],
                "_pasta_fotos": linha["_pasta_fotos"],
            }
            produtos_finais.append(produto)
            continue

        chave = calcular_chave_grupo(
            linha["titulo"], linha["color"], linha["_tipo_variacao"], linha["_variacao_valor"]
        )
        if chave in grupos_emitidos:
            continue
        grupos_emitidos.add(chave)

        linhas_do_grupo = grupos[chave]
        tipo_variacao = linhas_do_grupo[0]["_tipo_variacao"]

        if len(linhas_do_grupo) == 1:
            # "agrupável=sim" mas ficou sozinho no grupo (nenhuma outra
            # linha do Excel partilha a mesma base) — não faz sentido criar
            # uma lista de variações com 1 elemento; tratamos como produto
            # simples e avisamos, para o utilizador poder confirmar no Excel.
            avisos.append(
                f"WARNING: Produto '{linha['titulo']}' está marcado como "
                f"agrupável='sim' mas não foi encontrada nenhuma outra linha "
                f"com a mesma base de título — mantido como produto simples."
            )
            id_final = gerar_id(linha["titulo"])
            produto = {
                "id": id_final,
                "name": linha["titulo"],
                "description": linha["description"],
                "price": linha["price"],
                "brand": linha["brand"],
                "category": linha["category"],
                "color": linha["color"],
                "material": linha["material"],
                "additional_info": linha["additional_info"],
                "stock": linha["stock"],
                "weight_g": linha["weight_g"],
                "unit_count": linha["unit_count"],
                "images": [],
                "_pasta_fotos": linha["_pasta_fotos"],
            }
            produtos_finais.append(produto)
            continue

        nome_grupo = calcular_nome_grupo(
            linhas_do_grupo[0]["titulo"], linhas_do_grupo[0]["color"],
            tipo_variacao, linhas_do_grupo[0]["_variacao_valor"],
        )
        id_pai = gerar_id(nome_grupo)

        primeira = linhas_do_grupo[0]
        produto_pai = {
            "id": id_pai,
            "name": nome_grupo,
            "description": primeira["description"],
            "price": primeira["price"],
            "brand": primeira["brand"],
            "category": primeira["category"],
            "material": primeira["material"],
            "additional_info": primeira["additional_info"],
            "variation_type": tipo_variacao,
            "variations": [],
            "_pasta_fotos": None,  # produto-pai não tem fotos próprias
        }

        for linha_variacao in linhas_do_grupo:
            sku = gerar_id(f"{nome_grupo}-{linha_variacao['_variacao_valor']}")
            variacao_obj = {
                "sku": sku,
                "variacao": linha_variacao["_variacao_valor"],
                "preco": linha_variacao["price"],
                "stock": linha_variacao["stock"],
                "peso_kg": (
                    round(linha_variacao["weight_g"] / 1000, 3)
                    if linha_variacao["weight_g"] is not None else None
                ),
                "weight_g": linha_variacao["weight_g"],
                "unidades": linha_variacao["unit_count"],
                "color": linha_variacao["color"],
                "foto_principal": None,
                "galeria_fotos": [],
                "_pasta_fotos": linha_variacao["_pasta_fotos"],
            }
            produto_pai["variations"].append(variacao_obj)

        produtos_finais.append(produto_pai)

    return produtos_finais


def resolver_fotografias_produto_simples(produto, photos_base_dir: Path, output_images_dir: Path, avisos):
    pasta_relativa = produto.pop("_pasta_fotos", "")
    if not pasta_relativa:
        avisos.append(f"{produto['id']} → sem pasta de fotografias indicada no Excel.")
        return False

    pasta_origem = photos_base_dir / pasta_relativa.replace("/", os.sep)

    if not pasta_origem.is_dir():
        avisos.append(f"{produto['id']} → fotografia não encontrada (pasta '{pasta_relativa}' não existe).")
        return False

    ficheiros = sorted(
        p for p in pasta_origem.iterdir()
        if p.is_file() and p.suffix.lower() in EXTENSOES_IMAGEM
    )
    if not ficheiros:
        avisos.append(f"{produto['id']} → fotografia não encontrada (pasta '{pasta_relativa}' está vazia).")
        return False

    pasta_destino = output_images_dir / produto["id"]
    pasta_destino.mkdir(parents=True, exist_ok=True)

    imagens_finais = []
    for ficheiro in ficheiros:
        destino = pasta_destino / ficheiro.name
        shutil.copy2(ficheiro, destino)
        # caminho Web, sempre com barras normais — nunca "\"
        imagens_finais.append(f"images/produtos/{produto['id']}/{ficheiro.name}")

    produto["images"] = imagens_finais
    return True


def resolver_fotografias_variacao(produto_pai, variacao, photos_base_dir: Path, output_images_dir: Path, avisos):
    pasta_relativa = variacao.pop("_pasta_fotos", "")
    if not pasta_relativa:
        avisos.append(f"{produto_pai['id']} / {variacao['sku']} → sem pasta de fotografias indicada no Excel.")
        return False

    pasta_origem = photos_base_dir / pasta_relativa.replace("/", os.sep)

    if not pasta_origem.is_dir():
        avisos.append(f"{produto_pai['id']} / {variacao['sku']} → fotografia não encontrada (pasta '{pasta_relativa}' não existe).")
        return False

    ficheiros = sorted(
        p for p in pasta_origem.iterdir()
        if p.is_file() and p.suffix.lower() in EXTENSOES_IMAGEM
    )
    if not ficheiros:
        avisos.append(f"{produto_pai['id']} / {variacao['sku']} → fotografia não encontrada (pasta '{pasta_relativa}' está vazia).")
        return False

    pasta_destino = output_images_dir / produto_pai["id"] / variacao["sku"]
    pasta_destino.mkdir(parents=True, exist_ok=True)

    imagens_finais = []
    for ficheiro in ficheiros:
        destino = pasta_destino / ficheiro.name
        shutil.copy2(ficheiro, destino)
        imagens_finais.append(f"images/produtos/{produto_pai['id']}/{variacao['sku']}/{ficheiro.name}")

    variacao["galeria_fotos"] = imagens_finais
    variacao["foto_principal"] = imagens_finais[0] if imagens_finais else None
    return True


def resolver_todas_fotografias(produtos, photos_base_dir: Path, output_images_dir: Path, avisos):
    encontradas = 0
    total = 0
    for produto in produtos:
        if "variations" in produto:
            for variacao in produto["variations"]:
                total += 1
                if resolver_fotografias_variacao(produto, variacao, photos_base_dir, output_images_dir, avisos):
                    encontradas += 1
            # imagem do produto-pai (para os cards do catálogo) = foto
            # principal da primeira variação com stock/fotos válidas
            produto["images"] = next(
                (v["galeria_fotos"] for v in produto["variations"] if v["galeria_fotos"]),
                [],
            )
        else:
            total += 1
            if resolver_fotografias_produto_simples(produto, photos_base_dir, output_images_dir, avisos):
                encontradas += 1
    return encontradas, total


def comparar_com_versao_anterior(produtos_novos, output_json: Path):
    if not output_json.exists():
        return len(produtos_novos), 0, 0
    try:
        with open(output_json, "r", encoding="utf-8") as f:
            anterior = json.load(f)
        produtos_antigos = {p["id"]: p for p in anterior.get("products", [])}
    except (json.JSONDecodeError, KeyError):
        return len(produtos_novos), 0, 0

    ids_novos = {p["id"] for p in produtos_novos}
    novos = sum(1 for p in produtos_novos if p["id"] not in produtos_antigos)
    removidos = sum(1 for id_antigo in produtos_antigos if id_antigo not in ids_novos)
    atualizados = sum(
        1 for p in produtos_novos
        if p["id"] in produtos_antigos and produtos_antigos[p["id"]] != p
    )
    return novos, atualizados, removidos


def main():
    config = carregar_config()
    caminho_excel = Path(config["stock_excel_path"])
    photos_base_dir = Path(config["photos_base_dir"])
    nome_folha = config.get("sheet_name", "Rascunhos")
    output_json = BASE_DIR / config.get("output_json", "data/products.json")
    output_images_dir = BASE_DIR / config.get("output_images_dir", "images/produtos")

    print(f"Lendo {caminho_excel.name}...\n")
    verificar_ficheiro_disponivel_localmente(caminho_excel)

    linhas_lidas, avisos, erros = ler_produtos(caminho_excel, nome_folha)
    print(f"Linhas de produto encontradas: {len(linhas_lidas)}")

    produtos = agrupar_produtos(linhas_lidas, avisos)
    num_grupos = sum(1 for p in produtos if "variations" in p)
    num_simples = len(produtos) - num_grupos
    print(f"Produtos finais no catálogo: {len(produtos)} "
          f"({num_simples} simples + {num_grupos} agrupados com variações)")

    novos, atualizados, removidos = comparar_com_versao_anterior(produtos, output_json)

    fotos_encontradas, fotos_total = resolver_todas_fotografias(produtos, photos_base_dir, output_images_dir, avisos)

    # limpeza final: nenhum campo interno (prefixo "_") deve sobreviver no JSON
    def limpar(obj):
        if isinstance(obj, dict):
            return {k: limpar(v) for k, v in obj.items() if not k.startswith("_")}
        if isinstance(obj, list):
            return [limpar(v) for v in obj]
        return obj

    produtos_limpos = limpar(produtos)

    output_json.parent.mkdir(parents=True, exist_ok=True)
    with open(output_json, "w", encoding="utf-8") as f:
        json.dump({"products": produtos_limpos}, f, ensure_ascii=False, indent=2)

    print(f"\nProdutos novos: {novos}")
    print(f"Produtos atualizados: {atualizados}")
    print(f"Produtos removidos: {removidos}")
    print(f"\nFotografias encontradas: {fotos_encontradas}/{fotos_total}")

    if avisos:
        print("\nAVISOS:")
        for aviso in avisos:
            print(f"  {aviso}")

    if erros:
        print("\nERROS:")
        for erro in erros:
            print(f"  {erro}")

    print(f"\n{output_json.relative_to(BASE_DIR)} atualizado.")
    print("\nSincronização concluída.")


if __name__ == "__main__":
    main()
